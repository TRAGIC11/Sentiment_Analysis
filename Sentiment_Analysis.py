#!/usr/bin/env python
# coding: utf-8

# In[1]:


import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import numpy as np
import spacy
from textstat.textstat import textstatistics
import nltk
from nltk.corpus import stopwords,brown
from nltk.tokenize import word_tokenize
import openpyxl

nlp = spacy.load('en_core_web_sm')
nltk.download('wordnet')
nltk.download('stopwords')
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
nltk.download('tagsets')
stop_words = set(stopwords.words('english'))


# In[2]:


X = pd.read_excel('Input.xlsx')
X.head()


# In[3]:


headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',}
def get_title(url):
    page = requests.get(url,headers=headers)
    soup = BeautifulSoup(page.content, 'html.parser')
    title = soup.title.get_text()
    return title

def get_raw_content(url):
    page = requests.get(url,headers=headers)
    soup = BeautifulSoup(page.content, 'html.parser')
    raw_content = soup.body.find_all(class_ = "td-post-content")[0].get_text()
    return str(raw_content)

def remove_stopwords(text):
    text = re.sub("[^-9A-Za-z ]", " " , text)
    tokens = word_tokenize(text)
    filtered = [ w for w in tokens if not w.lower() in stop_words]
    for j,i in enumerate(filtered):
        i = re.sub('[^-9A-Za-z ]','',i)
        filtered[j] = i
    for i in range(len(filtered)-1,-1,-1):
        if filtered[i] == '':
            filtered.pop(i)
    return filtered

master_dict = pd.read_csv('Master_Dictionary.csv')
master_dict["Word"] = master_dict["Word"].str.lower()
master_dict = master_dict[['Word','Negative','Positive']]
master_dict['P/N'] = master_dict['Positive'] - master_dict['Negative']
master_dict = master_dict[master_dict['P/N'] != 0]
master_dict = master_dict.drop(columns = ['Positive','Negative'])
master_dict['P/N'] = np.where(master_dict['P/N'] > 0, 1, 0)

def count_PN(li):
    positive = negative = 0
    for i in li:
        if i in master_dict['Word'].values:
            if master_dict.loc[master_dict['Word'] == i,'P/N'].values[0] == 1:
                positive += 1
            else:
                negative += 1
    return positive,negative

def make_sentence(text):
    sentences = nlp(text)
    return list(sentences.sents)

def word_count(text):
    sentences = make_sentence(text)
    count = 0
    for i in sentences:
        count += len(word_tokenize(i.text))
    return count

def sentence_count(text):
    sentences = make_sentence(text)
    count = len(sentences)
    return count

def avg_sentence_length(text):
    words = word_count(text)
    sentences = sentence_count(text)
    average_length = words/sentences
    return average_length

def count_syllables(word):
    count = textstatistics().syllable_count(word)
    return count

def avg_syllables(text):
    syllables = count_syllables(text)
    words = word_count(text)
    count = syllables/words
    return count

def hard_words(text):
    sentences = nlp(text)
    words = []
    sentences = make_sentence(text)
    for i in sentences:
        words += [str(token) for token in i]
    
    hard_words = set()

    for i in words:
        syllable_count = count_syllables(i)
        if i not in nlp.Defaults.stop_words and syllable_count >= 2:
            hard_words.add(i)
        
    return hard_words

def average_word_len(tokens):
    length = 0
    for i in tokens:
        length += len(i)
    count = length/len(tokens)
    return count

def Pronoun_count(text):
    token = nltk.word_tokenize(text)
    list_of_pronouns = ["i", "we", "my", "ours", "our","us"]
    tagged_token =  nltk.pos_tag(token)
    pronoun_count = 0
    for word,tag in tagged_token:
        if word.lower() in list_of_pronouns and (tag=='PRP' or tag=='PRP$'):
            pronoun_count += 1
    return pronoun_count


# In[4]:


X['TITLE'] = X['URL'].apply(get_title)
X['Raw_Content'] = X['URL'].apply(get_raw_content)
X['Filtered_Content'] = X['Raw_Content'].apply(remove_stopwords)
X['URL_ID'] = X['URL_ID'].astype(int)
X.set_index('URL_ID',drop=True, inplace=True)
X.head()


# In[5]:


X['P/N'] = X['Filtered_Content'].apply(count_PN)
X['POSITIVE SCORE'] = X['P/N'].apply(lambda x: x[0])
X['NEGATIVE SCORE'] = X['P/N'].apply(lambda x: x[1])
X['POLARITY'] = (X['POSITIVE SCORE'] - X['NEGATIVE SCORE'])/(X['POSITIVE SCORE'] + X['NEGATIVE SCORE'] + 0.000001)
X['SUBJECTIVITY'] = (X['POSITIVE SCORE'] + X['NEGATIVE SCORE'])/(len(X['Filtered_Content']) + 0.000001)
X['AVG SENTENCE LENGTH'] = X['Raw_Content'].apply(avg_sentence_length)
X['Complex Words'] = X['Raw_Content'].apply(hard_words)
X['Word Count Total'] = X['Raw_Content'].apply(word_count)
X['PERCENTAGE OF COMPLEX WORDS'] = X['Complex Words'].apply(lambda x: len(x))/X['Word Count Total']*100
X['FOG INDEX'] = 0.4*X['AVG SENTENCE LENGTH'] + 0.4*X['PERCENTAGE OF COMPLEX WORDS']
X['Sentence Count'] = X['Raw_Content'].apply(sentence_count)
X['AVG NUMBER OF WORDS PER SENTENCE'] = X['Word Count Total']/X['Sentence Count']
X['COMPLEX WORD COUNT'] = X['Complex Words'].apply(len)
X['WORD COUNT'] = X['Filtered_Content'].apply(len)
X['SYLLABLE PER WORD'] = X['Raw_Content'].apply(avg_syllables)
X['PERSONAL PRONOUN'] = X['Raw_Content'].apply(Pronoun_count)
X['AVG WORD LENGTH'] = X['Filtered_Content'].apply(average_word_len)
X.drop(columns = ['Raw_Content','Filtered_Content','Complex Words','Word Count Total','Sentence Count','P/N'],inplace = True)
X.head()


# In[6]:


X.info()


# In[7]:


X = X.round(3)
X.head()


# In[8]:


wb = openpyxl.load_workbook(filename = 'D:\\Project\\Jupyter\\Blackcoffer\\Output Data Structure.xlsx')
ws = wb.active
for i in range(1,len(X)+1):
    ws.cell(row = i+1, column =  3).value = X.iloc[i-1]['POSITIVE SCORE']
    ws.cell(row = i+1, column =  4).value = X.iloc[i-1]['NEGATIVE SCORE']
    ws.cell(row = i+1, column =  5).value = X.iloc[i-1]['POLARITY']
    ws.cell(row = i+1, column =  6).value = X.iloc[i-1]['SUBJECTIVITY']
    ws.cell(row = i+1, column =  7).value = X.iloc[i-1]['AVG SENTENCE LENGTH']
    ws.cell(row = i+1, column =  8).value = X.iloc[i-1]['PERCENTAGE OF COMPLEX WORDS']
    ws.cell(row = i+1, column =  9).value = X.iloc[i-1]['FOG INDEX']
    ws.cell(row = i+1, column = 10).value = X.iloc[i-1]['AVG NUMBER OF WORDS PER SENTENCE']
    ws.cell(row = i+1, column = 11).value = X.iloc[i-1]['COMPLEX WORD COUNT']
    ws.cell(row = i+1, column = 12).value = X.iloc[i-1]['WORD COUNT']
    ws.cell(row = i+1, column = 13).value = X.iloc[i-1]['SYLLABLE PER WORD']
    ws.cell(row = i+1, column = 14).value = X.iloc[i-1]['PERSONAL PRONOUN']
    ws.cell(row = i+1, column = 15).value = X.iloc[i-1]['AVG WORD LENGTH']
wb.save('D:\\Project\\Jupyter\\Blackcoffer\\Output Data Structure.xlsx')

